import openpyxl
from openpyxl import Workbook


class Animal:
    def __init__(self, name, age):
        self.name = name
        self.age = age

    def make_sound(self):
        raise NotImplementedError("Этот метод должен быть переопределен в подклассах")

    def eat(self, food):
        print(f"{self.name} ест {food}")


class Mammal(Animal):
    def __init__(self, name, age, fur_color):
        super().__init__(name, age)
        self.fur_color = fur_color

    def make_sound(self):
        print(f"{self.name} издает звук")


class Bird(Animal):
    def __init__(self, name, age, wing_span):
        super().__init__(name, age)
        self.wing_span = wing_span

    def make_sound(self):
        print(f"{self.name} чирикает")

    def fly(self):
        print(f"{self.name} летит с размахом крыльев {self.wing_span} метров")


class Reptile(Animal):
    def __init__(self, name, age, scale_type):
        super().__init__(name, age)
        self.scale_type = scale_type

    def make_sound(self):
        print(f"{self.name} шипит")


class Employee:
    def __init__(self, name, experience, position):
        self.name = name
        self.experience = experience
        self.position = position

    def work(self):
        print(f"{self.name} работает на должности {self.position}")


def animal_sound(animals):
    for animal in animals:
        animal.make_sound()


class Zoo:
    def __init__(self, name):
        self.name = name
        self.animals = []
        self.employees = []

        self.load_employees()

    def load_employees(self):
        try:
            workbook = openpyxl.load_workbook('ZooPersonal.xlsx')
            sheet = workbook.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                name, position, experience = row
                self.employees.append(Employee(name, int(experience), position))
        except FileNotFoundError:
            print("Файл ZooPersonal.xlsx не найден, будет создан новый при сохранении данных.")

    def save_employees(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Имя", "Должность", "Стаж работы"])

        for employee in self.employees:
            sheet.append([employee.name, employee.position, employee.experience])

        workbook.save('ZooPersonal.xlsx')

    def add_animal(self, animal):
        self.animals.append(animal)
        print(f"{animal.name} добавлен в зоопарк")

    def add_employee(self, employee):
        self.employees.append(employee)
        print(f"{employee.name} добавлен в зоопарк как {employee.position}")

    def list_animals(self):
        print("Животные в зоопарке:")
        for animal in self.animals:
            print(f"- {animal.name}, возраст: {animal.age}")

    def list_employees(self):
        print("Сотрудники зоопарка:")
        for employee in self.employees:
            print(f"- {employee.name}, должность: {employee.position}, стаж: {employee.experience} лет")

    def add_employee_interactive(self):
        while True:
            name = input("Введите имя сотрудника: ")
            experience = input("Введите стаж работы сотрудника: ")
            while True:
                position = input("Введите специальность сотрудника (Директор, Дрессировщик, Технический работник): ")
                if position in ["Директор", "Дрессировщик", "Технический работник"]:
                    break
                else:
                    print("Некорректная специальность. Попробуйте снова.")

            new_employee = Employee(name, int(experience), position)
            self.add_employee(new_employee)

            cont = input("Закончить ввод сотрудников? (да/нет): ")
            if cont.lower() == 'да':
                break

    def add_animal_interactive(self):
        while True:
            animal_class = input("Введите класс животного (Млекопитающее, Птицы, Рептилия): ").lower()
            name = input("Введите имя животного: ")
            age = input("Введите возраст животного: ")

            if animal_class == 'млекопитающее':
                fur_color = input("Введите цвет шерсти: ")
                new_animal = Mammal(name, int(age), fur_color)
            elif animal_class == 'птицы':
                wing_span = float(input("Введите размах крыльев: "))
                new_animal = Bird(name, int(age), wing_span)
            elif animal_class == 'рептилия':
                scale_type = input("Введите тип чешуи: ")
                new_animal = Reptile(name, int(age), scale_type)
            else:
                print("Некорректный класс животного. Попробуйте снова.")
                continue

            self.add_animal(new_animal)

            cont = input("Закончить ввод животных? (да/нет): ")
            if cont.lower() == 'да':
                break


# Пример использования
def main():
    # Создание зоопарка
    my_zoo = Zoo("Зоопарк имени Ивана Ивановича")

    while True:
        print("\nМеню:")
        print("1. Показать список животных")
        print("2. Показать список сотрудников")
        print("3. Добавить сотрудников")
        print("4. Добавить животных")
        print("5. Выход")

        choice = input("Выберите действие: ")

        if choice == '1':
            my_zoo.list_animals()
        elif choice == '2':
            my_zoo.list_employees()
        elif choice == '3':
            my_zoo.add_employee_interactive()
        elif choice == '4':
            my_zoo.add_animal_interactive()
        elif choice == '5':
            my_zoo.save_employees()
            break
        else:
            print("Некорректный выбор. Попробуйте снова.")


if __name__ == "__main__":
    main()